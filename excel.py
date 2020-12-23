from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl import load_workbook

def get_info1(path):
    wb = load_workbook(path)
    ws = wb['info']
    info=[]
    for row in range(2, ws.max_row+1):
        brand = ws['A' + str(row)].value
        if brand:
            brand=brand.upper() #将“品牌”中的小写字母全部转换成大写字母
        typ = ws['B' + str(row)].value # 获取单元格中的数据
        pn = ws['C' + str(row)].value
        lotno = ws['D' + str(row)].value
        #if date
        date = ws['E' + str(row)].value
        if date:
            #print(date)
            date=date.date() #只获取日期时间中的日期，比如2019-2-20，不需要具体时间
        quantity = str(ws['F' + str(row)].value)+ "pcs" # 在数量后面加上“pcs”字样
        data={
            "brand":brand,
            "typ":typ,
            "pn":pn,
            "lotno":lotno,
            "date":date,
            "quantity":quantity
        }
        info.append(data)
    return info



def write_info1(path,info):
    wb = load_workbook(path)
    ws = wb['label']
    k=0
    for i in range(2,9,3): #列遍历
        for j in range(1,round(len(info)*7/3),7): #行遍历
            if k < len(info): #当数据条数不是3的整数倍时，计数器k会超出列表info的范围，后续代码只有在k < len(info)的情况下执行
                ws.cell(row=j, column=i).value = info[k]['brand']
                ws.cell(row=j, column=i-1).value = '品牌 \nBrand'

                ws.cell(row=j + 1, column=i).value = info[k]['typ']
                ws.cell(row=j + 1,column=i-1).value = '型号 Type'

                ws.cell(row=j + 2, column=i).value = info[k]['pn']
                ws.cell(row=j + 2,column=i-1).value = '物料编号 Item P/N'

                ws.cell(row=j + 3, column=i).value = info[k]['lotno']
                ws.cell(row=j + 3,column=i-1).value = '生产批号 Lot No.'

                ws.cell(row=j + 4, column=i).value = info[k]['date']
                ws.cell(row=j + 4,column=i-1).value = '生产日期 Date'

                ws.cell(row=j + 5, column=i).value = info[k]['quantity']
                ws.cell(row=j + 5,column=i-1).value = '数量 Quantity'
            k+=1 #k是列表info的索引，此处加一以便获取下一条数据
    wb.save(path) #保存Excel文件

def add_border1(path,worksheet):
    wb = load_workbook(path)
    ws = wb[worksheet]
    max_row = ws.max_row
    max_column = ws.max_column
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                       bottom=Side(style='thin'))

    for i in range(1,max_column +  1):
        for j in range(1,max_row + 1 ):
            if ws.cell(row=j,column=i).value:
                ws.cell(row=j, column=i).border = thin_border
    wb.save(path)


def clean_all(path,worksheet):
    None_border = Border(left=Side(style=None),
                         right=Side(style=None),
                         top=Side(style=None),
                       bottom=Side(style=None))

    wb = load_workbook(path)
    ws = wb[worksheet]
    max_row = ws.max_row
    max_column = ws.max_column
    for i in range(1,max_column +  1):
        for j in range(1,max_row + 1 ):
            ws.cell(row=j, column=i).value = None
            ws.cell(row=j, column=i).border = None_border

    wb.save(path)
